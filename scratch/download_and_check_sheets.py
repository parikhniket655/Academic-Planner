import urllib.request
import ssl
import openpyxl
from datetime import datetime

url = "https://docs.google.com/spreadsheets/d/1WPTKyFL52nR6PdJ2n2hdlw2yb-S5A6wiiHW1QBdUitU/export?format=xlsx"
context = ssl._create_unverified_context()
headers = {
    'User-Agent': 'Mozilla/5.0'
}

dest_path = "/Users/ketanparikh/Desktop/Antigravity Work/planner/scratch/live_schedule.xlsx"

try:
    print("Downloading live spreadsheet...")
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, context=context, timeout=20) as response:
        with open(dest_path, "wb") as f:
            f.write(response.read())
    print("Downloaded successfully!")
    
    wb = openpyxl.load_workbook(dest_path, data_only=True)
    print("Worksheets:", wb.sheetnames)
    
    # Try to find the Schedule sheet
    sheet = None
    for name in wb.sheetnames:
        if "Schedule" in name:
            sheet = wb[name]
            break
    if not sheet:
        # Fallback to first sheet
        sheet = wb.active
        
    print("Selected sheet:", sheet.title)
    
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
        
    print("Total rows read:", len(data))
    
    # Let's run the JS emulation parser in Python
    timeSlots = []
    for c in range(2, len(data[2])):
        val = str(data[2][c]).strip() if data[2][c] is not None else ""
        if val and any(char.isdigit() for char in val):
            timeSlots.append({"col": c, "slot": val.replace("-", " - ")})
            
    print("Time slots found:", [ts["slot"] for ts in timeSlots])
    
    courseAbbrMap = {}
    for r in range(2, len(data)):
        if len(data[r]) > 14 and data[r][13] is not None and data[r][14] is not None:
            fullName = str(data[r][13]).strip()
            abbr = str(data[r][14]).strip()
            if fullName and abbr and len(abbr) <= 10:
                courseAbbrMap[abbr] = fullName
                
    sessions = []
    currentDate = None
    currentDay = None
    lastValidDate = None
    
    for r in range(3, len(data)):
        row = data[r]
        dateVal = row[0]
        if dateVal is not None:
            if isinstance(dateVal, datetime):
                currentDate = dateVal
                currentDay = currentDate.strftime("%A")
                lastValidDate = datetime(currentDate.year, currentDate.month, currentDate.day)
            else:
                dateStr = str(dateVal).strip()
                if "SUNDAY" in dateStr.upper():
                    if lastValidDate:
                        currentDate = lastValidDate
                        # Emulate JS Sunday date shift
                        import datetime as dt
                        # In JS, getDay() !== 0 is true if it's Saturday (6), so it adds 1 day -> Sunday
                        # Let's check getDay() of currentDate
                        # Python weekday: Mon=0... Sun=6. JS: Sun=0, Mon=1... Sat=6.
                        # So JS getDay() is (python_weekday + 1) % 7.
                        js_day = (currentDate.weekday() + 1) % 7
                        if js_day != 0:
                            currentDate = currentDate + dt.timedelta(days=1)
                        currentDay = "Sunday"
                elif dateStr != "":
                    # try parse string
                    try:
                        currentDate = datetime.strptime(dateStr, "%d-%b-%y")
                        currentDay = currentDate.strftime("%A")
                        lastValidDate = datetime(currentDate.year, currentDate.month, currentDate.day)
                    except:
                        pass
                        
        if not currentDate:
            continue
            
        section = str(row[1]).strip().upper() if row[1] is not None else ""
        if not section or "SUNDAY" in section:
            continue
            
        dateKey = currentDate.strftime("%Y-%m-%d")
        
        for t in timeSlots:
            cellVal = str(row[t["col"]]).strip() if row[t["col"]] is not None else ""
            if not cellVal:
                continue
                
            import re
            match = re.match(r"^([A-Za-z0-9&\s\.\-]+?)\s*(\d+)\s*\(([^)]+)\)\s*$", cellVal)
            if match:
                courseCode = match.group(1).strip().rstrip("- ")
                sessionNum = match.group(2).strip()
                profInitials = match.group(3).strip()
                
                courseId = courseCode
                if courseCode in ["BA", "CB", "CV", "GBS", "SCM", "CW"]:
                    courseId = f"{courseCode} Sec-{section}"
                    
                sessions.append({
                    "dateKey": dateKey,
                    "slot": t["slot"],
                    "courseId": courseId,
                    "subject": courseAbbrMap.get(courseCode, courseCode),
                    "instructor": f"{profInitials}|{sessionNum}",
                    "section": section,
                    "row_num": r + 1
                })
                
    print("Total parsed sessions:", len(sessions))
    
    # Check for duplicates
    seen = {}
    duplicates = []
    for idx, s in enumerate(sessions):
        key = (s["dateKey"], s["slot"], s["courseId"])
        if key in seen:
            duplicates.append((key, seen[key], s))
        else:
            seen[key] = s
            
    print(f"Total duplicates found: {len(duplicates)}")
    for key, original, dup in duplicates[:10]:
        print(f"  Duplicate for {key}:")
        print(f"    Original: Row {original['row_num']} | Subject: {original['subject']} | Instructor: {original['instructor']}")
        print(f"    Duplicate: Row {dup['row_num']} | Subject: {dup['subject']} | Instructor: {dup['instructor']}")

except Exception as e:
    print("Error:", e)
